import json
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.styles import Alignment
from collections import defaultdict

file_path_url = os.getcwd() + "/backup/file_path.json"
with open(file_path_url) as file_path_json:
    file_path = json.load(file_path_json)
data_dir = file_path['datas_dir']


def set_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                cell_value = str(cell.value)
                length = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in cell_value)  # 中文字符计数为2，其他为1
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = (max_length + 2)  # 给长度加2为了更好的显示效果
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width


def set_cells_center_aligned(sheet):
    center_aligned = Alignment(horizontal='center', vertical='center')

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_aligned


def load_file(file_url):
    # 使用openpyxl加载工作簿
    workbook = load_workbook(file_url)
    sheet = workbook.active

    # 定义新表格的表头
    new_headers = ["序号", "发布时间", "昵称", "支付宝收款人", "支付宝收款账号", "报价", "笔记标题/链接", "点赞",
                   "收藏", "评论", "当前CPE",
                   "互动量", "财务是否支付"]
    old_headers = [cell.value for cell in sheet[2]]

    data_list = []

    # 遍历行，从旧文件中提取信息
    for row in sheet.iter_rows(min_row=3):
        row_data = []
        for header in new_headers:
            if header in ["序号", "财务是否支付", "当前CPE"]:
                row_data.append(None)
            else:
                index = old_headers.index(header)
                cell_value = row[index].value
                if header == "笔记标题/链接":
                    hyperlink = row[index].hyperlink
                    cell_value = (cell_value, hyperlink)  # 为超链接和文本创建元组
                elif isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y/%m/%d')
                row_data.append(cell_value)
        data_list.append(row_data)

    sorted_data_list = sorted(
        [row for row in data_list if row[new_headers.index("昵称")] and row[new_headers.index("昵称")] != ""],
        key=lambda x: x[new_headers.index("昵称")])
    print([new_headers] + sorted_data_list)
    return [new_headers] + sorted_data_list


def create_summary_sheet(data, workbook):
    # 创建新的工作表
    sheet = workbook.create_sheet("Summary")

    # 定义表头
    headers = ["序号", "收款方支付宝账号", "收款方姓名", "金额", "备注"]

    # 用字典记录每个支付宝账号对应的总金额
    account_info_dict = defaultdict(lambda: {'金额': 0, '支付宝收款人': '', '备注': '扇贝考研小红书合作'})

    for row in data[1:]:  # 跳过表头
        try:
            amount = float(row[5])  # 尝试将金额转换为浮点数
            account_info_dict[row[4]]['金额'] += amount  # 如果转换成功，进行累加
            account_info_dict[row[4]]['支付宝收款人'] = row[3]  # 记录收款方姓名
        except ValueError:
            pass  # 如果转换失败，忽略该行数据

    # 将字典转换为列表，并按照序号顺序排序
    account_info_list = list(account_info_dict.items())

    # 写入表头
    for j, header in enumerate(headers):
        sheet.cell(row=1, column=j + 1).value = header

    # 写入数据
    for i, info in enumerate(account_info_list, start=2):
        sheet.cell(row=i, column=1).value = i - 1  # 序号
        sheet.cell(row=i, column=2).value = info[0]  # 收款方支付宝账号
        sheet.cell(row=i, column=3).value = info[1]['支付宝收款人']  # 收款方姓名
        sheet.cell(row=i, column=4).value = info[1]['金额']  # 金额
        sheet.cell(row=i, column=5).value = info[1]['备注']  # 备注

    set_column_widths(sheet)
    set_cells_center_aligned(sheet)


# 修改 write_file 函数，以在写入数据后调用 create_summary_sheet 函数
def write_file(data, file_name):
    workbook = Workbook()
    sheet = workbook.active

    for i, row in enumerate(data):
        for j, cell_value in enumerate(row):
            if isinstance(cell_value, tuple):
                # 如果单元格值是元组，则第一个元素是文本，第二个元素是超链接
                sheet.cell(row=i + 1, column=j + 1).value = cell_value[0]
                sheet.cell(row=i + 1, column=j + 1).hyperlink = cell_value[1]
            else:
                sheet.cell(row=i + 1, column=j + 1).value = cell_value

    set_column_widths(sheet)
    set_cells_center_aligned(sheet)

    # 在保存之前创建Summary工作表
    create_summary_sheet(data, workbook)

    new_file_name = file_name.split(".")[0] + "_已处理.xlsx"
    output_path = os.path.join(data_dir, new_file_name)
    workbook.save(output_path)
    print(f"Data written successfully to '{output_path}'")