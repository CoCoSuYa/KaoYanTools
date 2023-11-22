import json
import os
import random
import re
import time
from datetime import datetime, timedelta
import chardet
import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

data_info_path = os.environ.get('data_info_file_path')
cookie_info_path = os.environ.get('cookie_info_file_path')
data_dir = os.environ.get('datas_dir')
xhs_cookie = ""
hide, collected, current_fans_num, current_nicker, start_of_week, end_of_week = 0, 0, 0, 0, 0, 0
headers = {"referer": "https://www.xiaohongshu.com/",
           "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "

                         "Chrome/114.0.0.0 Safari/537.36",
           "Accept": "application/json, text/plain, */*"}


def split_into_weeks(data_ids):
    """
        将数据按周分组
    :param data_ids:
    :return:
    """

    # 过滤掉空字符串或不完整的数据
    filtered_data_ids = [item for item in data_ids if item and isinstance(item, list) and len(item) > 1 and item[0]]

    print("filtered_data_ids:", filtered_data_ids)

    for data in filtered_data_ids:
        if data[0] == "None":
            filtered_data_ids.remove(data)

    # 按日期排序
    sorted_data_ids = sorted(filtered_data_ids, key=lambda x: datetime.strptime(x[0], "%Y-%m-%d"))

    print("sorted_data_ids", sorted_data_ids)

    sorted_date_objects = [(datetime.strptime(date_str[0], "%Y-%m-%d").date(), date_str[1]) for date_str in
                           sorted_data_ids]

    chunks = []
    current_sublist = [sorted_date_objects[0]]
    week_end_date = current_sublist[0][0] + timedelta(
        days=(6 - current_sublist[0][0].weekday()))  # Calculate the end of the week for the first date

    for date_obj, date_id in sorted_date_objects[1:]:
        if date_obj <= week_end_date:
            current_sublist.append((date_obj, date_id))
        else:
            chunks.append(current_sublist)
            current_sublist = [(date_obj, date_id)]
            week_end_date = date_obj + timedelta(days=(6 - date_obj.weekday()))

    if current_sublist:  # Append any remaining dates
        chunks.append(current_sublist)
    return chunks


def write_data_excel_file(data_ids):
    """
        将帖子数据写入Excel文件
    :param data_ids:
    """
    wb = Workbook()
    ws = wb.active
    try:
        title_list = ["发布时间", "昵称", "博主等级", "笔记标题/链接", "笔记形式", "互动量", "点赞数", "收藏数",
                      "评论数", "爆文情况", "备注"]

        # 填写标题
        for title in title_list:
            ws.cell(row=1, column=title_list.index(title) + 1, value=title)

        # 填写数据
        for row in range(len(data_ids)):
            row += 2
            for col in range(12):
                col += 1
                if data_ids[row - 2] == "None":
                    ws.cell(row=row, column=col, value="")
                else:
                    if col == 5:
                        ws.cell(row=row, column=col - 1).hyperlink = data_ids[row - 2][col - 1]
                    elif col > 5:
                        ws.cell(row=row, column=col - 1, value=data_ids[row - 2][col - 1])
                    else:
                        ws.cell(row=row, column=col, value=data_ids[row - 2][col - 1])

    except Exception as e:
        print("Error:\n", "填写数据表时出错:", str(e))
    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                # 增加对中文的处理，每个中文字符计为2个单位宽度
                cell_length = sum(2 if '\u4e00' <= ch <= '\u9fff' else 1 for ch in str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception as e:
                print("Error:\n", "设置数据表列宽时发生错误:", str(e))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
    with open(data_info_path, 'rb') as f:
        rawdata = f.read()
        result = chardet.detect(rawdata)
        char_en = result['encoding']
    # 得到原文件名（不含扩展名）
    with open(data_info_path, "r", encoding=char_en) as file_path:
        file_url = file_path.read()
    filename_without_ext = os.path.splitext(os.path.basename(file_url))[0]
    # 构造新文件名（可以自行修改格式）
    new_filename = filename_without_ext + "_数据整理.xlsx"
    # 得到新文件的完整路径
    new_file_url = os.path.join(data_dir, new_filename)
    print(new_file_url)
    # 使用os.path.abspath确保路径是绝对的
    absolute_path = os.path.abspath(new_file_url)
    print(absolute_path)
    wb.save(absolute_path)


def write_date_excel_file(data_ids):
    """
        将排期数据按周写入Excel文件
    :param data_ids:
    """
    global start_of_week, end_of_week
    for data in data_ids:
        print(data)
        if any(date_obj[0] == "None" for date_obj in data):
            continue
        wb = Workbook()
        ws = wb.active
        try:
            # 填写标题
            days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            for idx, day in enumerate(days):
                ws.cell(row=1, column=idx + 1, value=day)

            # 计算那个周的周日
            start_of_week = data[0][0] - timedelta(days=data[0][0].weekday())
            # 计算那个周的周六
            end_of_week = data[-1][0] + timedelta(days=6 - data[-1][0].weekday())

            # 填写第二行的日期范围
            current_day = start_of_week
            col = 1
            while current_day <= end_of_week:
                ws.cell(row=2, column=col, value=current_day.strftime('%Y-%m-%d'))
                current_day += timedelta(days=1)
                col += 1
            # 根据日期写入ID
            for date_obj, date_id in data:
                row = 3
                while ws.cell(row=row, column=date_obj.weekday() + 1).value:
                    row += 1
                ws.cell(row=row, column=date_obj.weekday() + 1, value=date_id)
        except Exception as e:
            print("Error:\n", "填写排期表时出错:", str(e))

        for col_num, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    # 增加对中文的处理，每个中文字符计为2个单位宽度
                    cell_length = sum(2 if '\u4e00' <= ch <= '\u9fff' else 1 for ch in str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception as e:
                    print("Error:\n", "设置排期表列宽时发生错误:", str(e))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        with open(data_info_path, 'rb') as f:
            rawdata = f.read()
            result = chardet.detect(rawdata)
            char_en = result['encoding']
        # 得到原文件名（不含扩展名）
        with open(data_info_path, "r", encoding=char_en) as file_path:
            file_url = file_path.read()
        filename_without_ext = os.path.splitext(os.path.basename(file_url))[0]
        # 构造新文件名（可以自行修改格式）
        new_filename = filename_without_ext + f'_排期表{start_of_week}~{end_of_week}.xlsx'
        # 得到新文件的完整路径
        new_file_url = os.path.join(data_dir, new_filename)
        print(new_file_url)
        # 使用os.path.abspath确保路径是绝对的
        absolute_path = os.path.abspath(new_file_url)
        print(absolute_path)
        wb.save(absolute_path)


def write_up_fans_excel_file(data_ids):
    """
        将博主数据写入Excel文件
    :param data_ids:
    """
    wb = Workbook()
    ws = wb.active
    title1 = ["博主", "粉丝数", "备注"]
    ws.cell(1, 1, title1[0])
    ws.cell(1, 2, title1[1])
    ws.cell(1, 3, title1[2])
    row = 2
    for data in data_ids:
        ws.cell(row, 1, data[0])
        ws.cell(row, 2, data[1])
        ws.cell(row, 3, data[2])
        row += 1

    for col_num, col in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col:
            try:
                # 增加对中文的处理，每个中文字符计为2个单位宽度
                cell_length = sum(2 if '\u4e00' <= ch <= '\u9fff' else 1 for ch in str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception as e:
                print("Error:\n", "设置博主表列宽时发生错误：", str(e))

        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
    with open(data_info_path, 'rb') as f:
        rawdata = f.read()
        result = chardet.detect(rawdata)
        char_en = result['encoding']
    # # 得到文件名（不含扩展名）
    with open(data_info_path, "r", encoding=char_en) as file_path:
        file_url = file_path.read()
    filename_without_ext = os.path.splitext(os.path.basename(file_url))[0]
    # 构造新文件名（可以自行修改格式）
    new_filename = filename_without_ext + "_粉丝收集.xlsx"
    # 得到新文件的完整路径
    new_file_url = os.path.join(data_dir, new_filename)
    print(new_file_url)
    # 使用os.path.abspath确保路径是绝对的
    absolute_path = os.path.abspath(new_file_url)
    print(absolute_path)
    wb.save(absolute_path)


def load_excel_file():
    """
        读取Excel文件
    :return: 读取到的超链接列表
    """
    links = []
    with open(data_info_path, 'r') as data_path:
        path = data_path.read().strip()
    print("excel file path:", path)
    try:
        wb = load_workbook(filename=path)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                link = cell.hyperlink.target if cell.hyperlink else None
                if link:
                    # 使用文件名作为键，并添加超链接到列表中
                    links.append(link)
                else:
                    links.append("")
    except Exception as e:
        print("Error:\n", "读取文件时出错:", str(e))
    return links


def load_json_file():
    """
        读取cookie文件
    """
    try:
        with open(cookie_info_path, 'r') as cookie_file_path:
            path = cookie_file_path.read().strip()
        print("cookie file path:", path)
        with open(path, 'r') as cookie_file:
            cookie_datas = json.load(cookie_file)
        global xhs_cookie
        xhs_cookie = cookie_datas['cookie1']
    except Exception as e:
        print(e)
        print("读取cookie文件时出错：cookie文件失效或缺失")


def get_redirected_url(url):
    """
        获取重定向后的url
    :param url: 原链接
    :return: 新链接
    """
    res = requests.get(url, headers=headers, cookies=xhs_cookie, allow_redirects=True)
    time.sleep(random.randint(1, 3))
    if res.history:  # 检查是否有重定向历史
        print(res.url)
        return res.url
    else:
        print(url)
        return url


def get_note_ids_from_links(links):
    """
        从链接中获取用户id
    :param links: 链接列表
    """
    note_ids = []
    for link in links:
        try:
            if link == "":
                note_ids.append("None")
                print("第", links.index(link) + 1, "条帖子id:", "None")
            else:
                url = get_redirected_url(link)
                if "item" in url:
                    note_id = re.findall(r'item/(\w+)', url)[0]
                elif "website-login" in url:
                    if "item" in url:
                        note_id = re.findall(r'item%2F(\w+)', url)[0]
                    else:
                        note_id = re.findall(r'explore%2F(\w+)', url)[0]
                else:
                    note_id = re.findall(r'explore/(\w+)', url)[0]
                note_ids.append(note_id)
                print("第", links.index(link) + 1, "条帖子id:", note_id)
        except Exception as e:
            print("Error:\n", f" 第{links.index(link) + 1}条有问题,问题原因: {str(e)},将跳过该条")
            note_ids.append("None")
    return note_ids


def extract_text(text):
    pattern = r"(?<=window.__INITIAL_STATE__=).*(?=</script>)"
    match = re.search(pattern, text, re.DOTALL)
    if match:
        result = match.group(0)
        end = result.find('</script>')  # find the first '</script>'
        if end != -1:
            result = result[:end].strip()
        return result
    else:
        return None


def get_nicker_level_by_user_id(user_id):
    """
            获取博主等级
        :param user_id:
    """
    url = f"https://www.xiaohongshu.com/user/profile/{user_id}"
    res = requests.get(url, headers=headers, cookies=xhs_cookie)
    time.sleep(random.randint(1, 4))
    if res.status_code == 200:
        initial_state_str = extract_text(res.text)
        initial_state_str = initial_state_str.replace('undefined', '"undefined"')
        initial_state = json.loads(initial_state_str)
        nicker_datas = initial_state["user"]["userPageData"]['interactions']
        print(nicker_datas)
        for nicker_data in nicker_datas:
            if nicker_data["type"] == "fans":
                fans_num = nicker_data["count"]
                if 0 < int(fans_num) < 3000:
                    return "新兴"
                elif 3000 <= int(fans_num) < 5000:
                    return "普通"
                elif 5000 <= int(fans_num) < 50000:
                    return "初级"
                elif 50000 <= int(fans_num) < 500000:
                    return "腰部"
                elif int(fans_num) >= 500000:
                    return "头部"
    else:
        return "博主已注销"


def get_note_data(note_ids):
    data_ids = []
    for note_id in note_ids:
        success = False
        retry_count = 0
        while not success:
            print("开始处理第", note_ids.index(note_id) + 1, "条数据")
            if note_id == "None":
                data_ids.append("None")
                print("链接", note_ids.index(note_id) + 1, "有问题,置为None")
                success = True
                continue
            res = requests.get(f"https://www.xiaohongshu.com/explore/{note_id}", headers=headers, cookies=xhs_cookie)
            if res.status_code == 200:
                try:
                    initial_state_str = extract_text(res.text)
                    initial_state_str = initial_state_str.replace('undefined', '"undefined"')
                    initial_state = json.loads(initial_state_str)
                    timestamp_ms = int(initial_state["note"]["noteDetailMap"][note_id]["note"]["time"])
                    timestamp_s = timestamp_ms / 1000  # Convert to seconds
                    date = datetime.fromtimestamp(timestamp_s)
                    formatted_date = date.strftime('%Y-%m-%d')
                    nicker_name = initial_state["note"]["noteDetailMap"][note_id]["note"]["user"]["nickname"]
                    nicker_id = initial_state["note"]["noteDetailMap"][note_id]["note"]["user"]["userId"]
                    time.sleep(random.randint(4, 6))
                    nicker_level = get_nicker_level_by_user_id(nicker_id)
                    if initial_state["note"]["noteDetailMap"][note_id]["note"]["title"] != "":
                        note_title = initial_state["note"]["noteDetailMap"][note_id]["note"]["title"]
                    else:
                        note_title = initial_state["note"]["noteDetailMap"][note_id]["note"]["desc"][0:30]
                    if initial_state["note"]["noteDetailMap"][note_id]["note"]["type"] == "video":
                        note_type = "视频"
                    else:
                        note_type = "图文"
                    print(initial_state["note"]["noteDetailMap"][note_id]["note"]["interactInfo"])
                    like_num = initial_state["note"]["noteDetailMap"][note_id]["note"]["interactInfo"]["likedCount"]
                    collect_num = initial_state["note"]["noteDetailMap"][note_id]["note"]["interactInfo"][
                        "collectedCount"]
                    comment_num = initial_state["note"]["noteDetailMap"][note_id]["note"]["interactInfo"][
                        "commentCount"]
                    interact_num = int(like_num) + int(collect_num) + int(comment_num)
                    if 0 <= interact_num < 1000:
                        note_level = ""
                    elif 1000 <= interact_num < 5000:
                        note_level = "小爆文"
                    elif 5000 <= interact_num < 10000:
                        note_level = "中爆文"
                    elif 10000 <= interact_num:
                        note_level = "大爆文"
                    else:
                        note_level = "互动量异常"
                    data_ids.append([formatted_date, nicker_name, nicker_level, note_title,
                                     f"https://www.xiaohongshu.com/explore/{note_id}", note_type, interact_num,
                                     like_num,
                                     collect_num, comment_num, note_level, "帖子正常"])
                    print([formatted_date, nicker_name, nicker_level, note_title,
                           f"https://www.xiaohongshu.com/explore/{note_id}", note_type, interact_num, like_num,
                           collect_num, comment_num, note_level, "帖子正常"])
                    print("第", note_ids.index(note_id) + 1, "条数据处理完成")
                    success = True
                except Exception as e:
                    print(e)
                    print("第", note_ids.index(note_id) + 1, "条数据处理失败,将重试")
                    if retry_count >= 5:
                        print("重试次数已达上限,跳过该条数据")
                        data_ids.append(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                                         "", "", "", "", "", "", "帖子已跳过"])
                        print(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                               "", "", "", "", "", "", "帖子已跳过"])
                        print("第", note_ids.index(note_id) + 1, "条数据处理完成")
                        retry_count = 0
                        success = True
                    else:
                        time.sleep(random.randint(1, 4))
                        retry_count += 1
                        print("重试次数:", retry_count)
            elif res.status_code == 404:
                data_ids.append(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                                 "", "", "", "", "", "", "帖子已被删除"])
                print(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                       "", "", "", "", "", "", "帖子已被删除"])
                print("第", note_ids.index(note_id) + 1, "条数据处理完成")
                success = True
            elif res.status_code == 403:
                data_ids.append(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                                 "", "", "", "", "", "", "帖子已被隐藏"])
                print(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                       "", "", "", "", "", "", "帖子已被隐藏"])
                print("第", note_ids.index(note_id) + 1, "条数据处理完成")
                success = True
            elif res.status_code == 423:
                data_ids.append(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                                 "", "", "", "", "", "", "帖子被官方锁定"])
                print(["", "", "", "", f"https://www.xiaohongshu.com/explore/{note_id}",
                       "", "", "", "", "", "", "帖子被官方锁定"])
                print("第", note_ids.index(note_id) + 1, "条数据处理完成")
                success = True
            else:
                print("cookie失效,获取cookie后重试")
                success = True

    return data_ids
